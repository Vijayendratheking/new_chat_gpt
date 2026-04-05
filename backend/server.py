from fastapi import FastAPI, APIRouter, UploadFile, File, Body
from fastapi.responses import StreamingResponse
from dotenv import load_dotenv
from starlette.middleware.cors import CORSMiddleware
from motor.motor_asyncio import AsyncIOMotorClient
import os
import logging
import io
import json
from pathlib import Path
from typing import Optional, List
import uuid
from datetime import datetime, timezone
import openpyxl
import xlsxwriter

from scheduler import (
    run_scheduler, DAYS, ENGLISH_SHIFTS, LANGUAGE_SHIFTS, ALL_SHIFTS,
    DEFAULT_OFF_DAY_PROFILES,
)

ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / '.env')

mongo_url = os.environ['MONGO_URL']
client = AsyncIOMotorClient(mongo_url)
db = client[os.environ['DB_NAME']]

app = FastAPI()
api_router = APIRouter(prefix="/api")

# Default requirement data from the uploaded image
DEFAULT_ENGLISH = {
    "Monday":    {"00:00":7,"01:00":4,"02:00":3,"03:00":3,"04:00":3,"05:00":3,"06:00":7,"07:00":16,"08:00":26,"09:00":33,"10:00":38,"11:00":39,"12:00":39,"13:00":39,"14:00":38,"15:00":37,"16:00":38,"17:00":36,"18:00":36,"19:00":35,"20:00":32,"21:00":28,"22:00":20,"23:00":13},
    "Tuesday":   {"00:00":8,"01:00":4,"02:00":3,"03:00":3,"04:00":3,"05:00":3,"06:00":7,"07:00":20,"08:00":26,"09:00":32,"10:00":33,"11:00":34,"12:00":34,"13:00":35,"14:00":35,"15:00":35,"16:00":35,"17:00":34,"18:00":32,"19:00":32,"20:00":29,"21:00":26,"22:00":21,"23:00":13},
    "Wednesday": {"00:00":8,"01:00":5,"02:00":3,"03:00":3,"04:00":3,"05:00":3,"06:00":8,"07:00":19,"08:00":25,"09:00":32,"10:00":35,"11:00":36,"12:00":37,"13:00":38,"14:00":38,"15:00":37,"16:00":36,"17:00":34,"18:00":32,"19:00":31,"20:00":30,"21:00":27,"22:00":23,"23:00":14},
    "Thursday":  {"00:00":7,"01:00":5,"02:00":3,"03:00":3,"04:00":3,"05:00":3,"06:00":8,"07:00":17,"08:00":0,"09:00":28,"10:00":30,"11:00":31,"12:00":32,"13:00":33,"14:00":33,"15:00":33,"16:00":32,"17:00":30,"18:00":29,"19:00":28,"20:00":27,"21:00":24,"22:00":18,"23:00":13},
    "Friday":    {"00:00":8,"01:00":5,"02:00":3,"03:00":3,"04:00":3,"05:00":3,"06:00":8,"07:00":18,"08:00":24,"09:00":29,"10:00":32,"11:00":32,"12:00":33,"13:00":34,"14:00":33,"15:00":33,"16:00":32,"17:00":30,"18:00":29,"19:00":26,"20:00":23,"21:00":21,"22:00":15,"23:00":11},
    "Saturday":  {"00:00":6,"01:00":4,"02:00":2,"03:00":2,"04:00":2,"05:00":2,"06:00":5,"07:00":15,"08:00":20,"09:00":25,"10:00":27,"11:00":27,"12:00":27,"13:00":28,"14:00":26,"15:00":25,"16:00":25,"17:00":23,"18:00":20,"19:00":18,"20:00":17,"21:00":14,"22:00":12,"23:00":9},
    "Sunday":    {"00:00":5,"01:00":2,"02:00":2,"03:00":2,"04:00":2,"05:00":2,"06:00":3,"07:00":10,"08:00":17,"09:00":20,"10:00":21,"11:00":22,"12:00":22,"13:00":23,"14:00":23,"15:00":22,"16:00":21,"17:00":20,"18:00":20,"19:00":19,"20:00":18,"21:00":16,"22:00":12,"23:00":8},
}

DEFAULT_LANGUAGE = {
    "Monday":    {"00:00":0,"01:00":0,"02:00":0,"03:00":0,"04:00":0,"05:00":0,"06:00":0,"07:00":43,"08:00":53,"09:00":60,"10:00":60,"11:00":60,"12:00":59,"13:00":55,"14:00":51,"15:00":50,"16:00":50,"17:00":50,"18:00":41,"19:00":17,"20:00":0,"21:00":0,"22:00":0,"23:00":0},
    "Tuesday":   {"00:00":0,"01:00":0,"02:00":0,"03:00":0,"04:00":0,"05:00":0,"06:00":0,"07:00":43,"08:00":50,"09:00":52,"10:00":53,"11:00":53,"12:00":51,"13:00":49,"14:00":49,"15:00":49,"16:00":47,"17:00":46,"18:00":40,"19:00":18,"20:00":0,"21:00":0,"22:00":0,"23:00":0},
    "Wednesday": {"00:00":0,"01:00":0,"02:00":0,"03:00":0,"04:00":0,"05:00":0,"06:00":0,"07:00":41,"08:00":51,"09:00":54,"10:00":55,"11:00":55,"12:00":52,"13:00":49,"14:00":49,"15:00":48,"16:00":47,"17:00":43,"18:00":34,"19:00":15,"20:00":0,"21:00":0,"22:00":0,"23:00":0},
    "Thursday":  {"00:00":0,"01:00":0,"02:00":0,"03:00":0,"04:00":0,"05:00":0,"06:00":0,"07:00":40,"08:00":47,"09:00":49,"10:00":49,"11:00":48,"12:00":48,"13:00":47,"14:00":46,"15:00":46,"16:00":45,"17:00":42,"18:00":35,"19:00":18,"20:00":0,"21:00":0,"22:00":0,"23:00":0},
    "Friday":    {"00:00":0,"01:00":0,"02:00":0,"03:00":0,"04:00":0,"05:00":0,"06:00":0,"07:00":36,"08:00":48,"09:00":49,"10:00":49,"11:00":49,"12:00":48,"13:00":45,"14:00":43,"15:00":42,"16:00":39,"17:00":35,"18:00":28,"19:00":13,"20:00":0,"21:00":0,"22:00":0,"23:00":0},
    "Saturday":  {"00:00":0,"01:00":0,"02:00":0,"03:00":0,"04:00":0,"05:00":0,"06:00":0,"07:00":0,"08:00":41,"09:00":43,"10:00":43,"11:00":42,"12:00":39,"13:00":35,"14:00":31,"15:00":29,"16:00":14,"17:00":0,"18:00":0,"19:00":0,"20:00":0,"21:00":0,"22:00":0,"23:00":0},
    "Sunday":    {"00:00":0,"01:00":0,"02:00":0,"03:00":0,"04:00":0,"05:00":0,"06:00":0,"07:00":0,"08:00":24,"09:00":25,"10:00":26,"11:00":24,"12:00":23,"13:00":23,"14:00":23,"15:00":20,"16:00":10,"17:00":0,"18:00":0,"19:00":0,"20:00":0,"21:00":0,"22:00":0,"23:00":0},
}


def parse_excel_to_dict(content: bytes) -> dict:
    """Parse an Excel file (first sheet) into {day: {hour: value}}.
    Handles title rows by scanning for the header row containing 'Interval'."""
    wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return {}
    # Find the header row (contains "Interval")
    header_idx = 0
    for i, row in enumerate(rows):
        cells = [str(c).strip().lower() if c else "" for c in row]
        if "interval" in cells:
            header_idx = i
            break
    headers = [str(h).strip() if h else "" for h in rows[header_idx]]
    result = {d: {} for d in DAYS}
    for row in rows[header_idx + 1:]:
        interval = str(row[0]).strip() if row[0] is not None else ""
        if not interval or ":" not in interval:
            continue
        for day in DAYS:
            if day in headers:
                day_col = headers.index(day)
                val = row[day_col] if day_col < len(row) else 0
                try:
                    result[day][interval] = float(val) if val is not None else 0
                except (ValueError, TypeError):
                    result[day][interval] = 0
    return result


def generate_sample_template() -> bytes:
    """Generate a sample Excel template with two sheets: English and Language."""
    output = io.BytesIO()
    wb = xlsxwriter.Workbook(output)

    # Formats
    header_fmt = wb.add_format({
        "bold": True, "bg_color": "#0033CC", "font_color": "#FFFFFF",
        "border": 1, "font_name": "Calibri", "font_size": 11,
        "align": "center", "valign": "vcenter",
    })
    interval_fmt = wb.add_format({
        "bold": True, "bg_color": "#F1F5F9", "border": 1,
        "font_name": "Calibri", "font_size": 11,
    })
    data_fmt = wb.add_format({
        "border": 1, "font_name": "Calibri", "font_size": 11,
        "align": "center", "num_format": "0",
    })
    title_fmt = wb.add_format({
        "bold": True, "font_size": 14, "font_name": "Calibri",
        "font_color": "#0033CC",
    })

    hours = [f"{h:02d}:00" for h in range(24)]

    for sheet_name, data in [("English", DEFAULT_ENGLISH), ("Language", DEFAULT_LANGUAGE)]:
        ws = wb.add_worksheet(sheet_name)
        ws.set_column("A:A", 10)
        ws.set_column("B:H", 12)

        # Title row
        ws.merge_range("A1:H1", f"Required Hours - {sheet_name}", title_fmt)

        # Headers
        ws.write(1, 0, "Interval", header_fmt)
        for i, day in enumerate(DAYS):
            ws.write(1, i + 1, day, header_fmt)

        # Data
        for row_idx, hour in enumerate(hours):
            ws.write(row_idx + 2, 0, hour, interval_fmt)
            for col_idx, day in enumerate(DAYS):
                val = data.get(day, {}).get(hour, 0)
                ws.write(row_idx + 2, col_idx + 1, val, data_fmt)

    # Add shift reference sheet
    ws_ref = wb.add_worksheet("Shift Reference")
    ws_ref.set_column("A:D", 18)
    ws_ref.write(0, 0, "Shift Reference", title_fmt)

    ws_ref.write(2, 0, "English Shifts", wb.add_format({"bold": True, "font_size": 12, "font_color": "#0033CC"}))
    ws_ref.write(3, 0, "Shift ID", header_fmt)
    ws_ref.write(3, 1, "Time Range", header_fmt)
    ws_ref.write(3, 2, "Duration", header_fmt)
    for i, s in enumerate(ENGLISH_SHIFTS):
        ws_ref.write(4 + i, 0, s["id"], data_fmt)
        ws_ref.write(4 + i, 1, s["label"], data_fmt)
        ws_ref.write(4 + i, 2, "9 hours", data_fmt)

    row_start = 4 + len(ENGLISH_SHIFTS) + 1
    ws_ref.write(row_start, 0, "Language Shifts", wb.add_format({"bold": True, "font_size": 12, "font_color": "#FF6600"}))
    ws_ref.write(row_start + 1, 0, "Shift ID", header_fmt)
    ws_ref.write(row_start + 1, 1, "Time Range", header_fmt)
    ws_ref.write(row_start + 1, 2, "Duration", header_fmt)
    for i, s in enumerate(LANGUAGE_SHIFTS):
        ws_ref.write(row_start + 2 + i, 0, s["id"], data_fmt)
        ws_ref.write(row_start + 2 + i, 1, s["label"], data_fmt)
        ws_ref.write(row_start + 2 + i, 2, "9 hours", data_fmt)

    wb.close()
    output.seek(0)
    return output.getvalue()


@api_router.get("/")
async def root():
    return {"message": "Cross-Skill Scheduler API"}


@api_router.get("/default-requirements")
async def get_default_requirements():
    return {"english": DEFAULT_ENGLISH, "language": DEFAULT_LANGUAGE}


@api_router.get("/sample-template")
async def download_sample_template():
    """Download a sample Excel template showing the required data format."""
    content = generate_sample_template()
    return StreamingResponse(
        io.BytesIO(content),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=schedule_template.xlsx"}
    )


def parse_multi_sheet_excel(content: bytes):
    """Parse a multi-sheet Excel file. Returns (english_dict, language_dict).
    Looks for sheets named 'English' and 'Language' (case-insensitive)."""
    wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
    sheet_map = {s.lower(): s for s in wb.sheetnames}

    eng_data = None
    lang_data = None

    for key, actual in sheet_map.items():
        if "english" in key or "eng" in key:
            eng_data = _parse_sheet(wb[actual])
        elif "language" in key or "lang" in key:
            lang_data = _parse_sheet(wb[actual])

    # If we couldn't match by name, use first two sheets
    if eng_data is None and lang_data is None and len(wb.sheetnames) >= 2:
        eng_data = _parse_sheet(wb[wb.sheetnames[0]])
        lang_data = _parse_sheet(wb[wb.sheetnames[1]])
    elif eng_data is None and lang_data is None and len(wb.sheetnames) == 1:
        eng_data = _parse_sheet(wb[wb.sheetnames[0]])

    return eng_data, lang_data


def _parse_sheet(ws) -> dict:
    """Parse a single worksheet into {day: {hour: value}}."""
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return {}
    header_idx = 0
    for i, row in enumerate(rows):
        cells = [str(c).strip().lower() if c else "" for c in row]
        if "interval" in cells:
            header_idx = i
            break
    headers = [str(h).strip() if h else "" for h in rows[header_idx]]
    result = {d: {} for d in DAYS}
    for row in rows[header_idx + 1:]:
        interval = str(row[0]).strip() if row[0] is not None else ""
        if not interval or ":" not in interval:
            continue
        for day in DAYS:
            if day in headers:
                day_col = headers.index(day)
                val = row[day_col] if day_col < len(row) else 0
                try:
                    result[day][interval] = float(val) if val is not None else 0
                except (ValueError, TypeError):
                    result[day][interval] = 0
    return result


@api_router.post("/run-schedule")
async def run_schedule_endpoint(
    combined_file: Optional[UploadFile] = File(None),
    english_file: Optional[UploadFile] = File(None),
    language_file: Optional[UploadFile] = File(None),
):
    """Run the scheduling algorithm.
    - combined_file: Single .xlsx with 'English' and 'Language' sheets
    - english_file / language_file: Separate files (.xlsx or .csv)
    - Falls back to defaults if nothing uploaded.
    """
    english_data = None
    language_data = None

    # Priority 1: Combined multi-sheet file
    if combined_file and combined_file.filename:
        combo_bytes = await combined_file.read()
        if combined_file.filename.endswith(".xlsx"):
            english_data, language_data = parse_multi_sheet_excel(combo_bytes)

    # Priority 2: Individual files (override anything not found above)
    if english_file and english_file.filename:
        eng_bytes = await english_file.read()
        if english_file.filename.endswith(".xlsx"):
            english_data = parse_excel_to_dict(eng_bytes)
        else:
            english_data = parse_csv_to_dict(eng_bytes.decode("utf-8"))

    if language_file and language_file.filename:
        lang_bytes = await language_file.read()
        if language_file.filename.endswith(".xlsx"):
            language_data = parse_excel_to_dict(lang_bytes)
        else:
            language_data = parse_csv_to_dict(lang_bytes.decode("utf-8"))

    # Fallback to defaults
    if not english_data:
        english_data = DEFAULT_ENGLISH
    if not language_data:
        language_data = DEFAULT_LANGUAGE

    result = run_scheduler(english_data, language_data)

    schedule_id = str(uuid.uuid4())
    doc = {
        "id": schedule_id,
        "name": "Default Schedule",
        "timestamp": datetime.now(timezone.utc).isoformat(),
        "english_input": english_data,
        "language_input": language_data,
        "off_day_profiles": [{"off_days": p["off_days"], "count": p["count"]} for p in DEFAULT_OFF_DAY_PROFILES],
        "shiftwise": result["shiftwise"],
        "gap_analysis": result["gap_analysis"],
        "roster": result["roster"],
        "sla": result["sla"],
        "summary": result["summary"],
    }
    await db.schedules.insert_one(doc)

    return {
        "id": schedule_id,
        "name": "Default Schedule",
        "shiftwise": result["shiftwise"],
        "gap_analysis": result["gap_analysis"],
        "roster": result["roster"],
        "sla": result["sla"],
        "summary": result["summary"],
    }


@api_router.post("/run-scenario")
async def run_scenario_endpoint(
    name: str = Body("Scenario"),
    off_day_profiles: str = Body(None),
    combined_file: Optional[UploadFile] = File(None),
    english_file: Optional[UploadFile] = File(None),
    language_file: Optional[UploadFile] = File(None),
):
    """Run a named scenario with custom off-day profiles."""
    english_data = None
    language_data = None

    if combined_file and combined_file.filename:
        combo_bytes = await combined_file.read()
        if combined_file.filename.endswith(".xlsx"):
            english_data, language_data = parse_multi_sheet_excel(combo_bytes)

    if english_file and english_file.filename:
        eng_bytes = await english_file.read()
        if english_file.filename.endswith(".xlsx"):
            english_data = parse_excel_to_dict(eng_bytes)
        else:
            english_data = parse_csv_to_dict(eng_bytes.decode("utf-8"))

    if language_file and language_file.filename:
        lang_bytes = await language_file.read()
        if language_file.filename.endswith(".xlsx"):
            language_data = parse_excel_to_dict(lang_bytes)
        else:
            language_data = parse_csv_to_dict(lang_bytes.decode("utf-8"))

    if not english_data:
        english_data = DEFAULT_ENGLISH
    if not language_data:
        language_data = DEFAULT_LANGUAGE

    # Parse custom off-day profiles
    profiles = None
    if off_day_profiles:
        try:
            profiles = json.loads(off_day_profiles)
        except (json.JSONDecodeError, TypeError):
            profiles = None

    result = run_scheduler(english_data, language_data, off_day_profiles=profiles)

    schedule_id = str(uuid.uuid4())
    stored_profiles = profiles if profiles else [{"off_days": p["off_days"], "count": p["count"]} for p in DEFAULT_OFF_DAY_PROFILES]
    doc = {
        "id": schedule_id,
        "name": name,
        "timestamp": datetime.now(timezone.utc).isoformat(),
        "english_input": english_data,
        "language_input": language_data,
        "off_day_profiles": stored_profiles,
        "shiftwise": result["shiftwise"],
        "gap_analysis": result["gap_analysis"],
        "roster": result["roster"],
        "sla": result["sla"],
        "summary": result["summary"],
    }
    await db.schedules.insert_one(doc)

    return {
        "id": schedule_id,
        "name": name,
        "off_day_profiles": stored_profiles,
        "shiftwise": result["shiftwise"],
        "gap_analysis": result["gap_analysis"],
        "roster": result["roster"],
        "sla": result["sla"],
        "summary": result["summary"],
    }


@api_router.post("/compare")
async def compare_scenarios(payload: dict = Body(...)):
    """Return comparison data for multiple schedule IDs."""
    ids = payload.get("ids", [])
    if len(ids) < 2:
        return {"error": "Need at least 2 schedule IDs to compare"}

    scenarios = []
    for sid in ids[:5]:  # max 5
        doc = await db.schedules.find_one({"id": sid}, {"_id": 0, "roster": 0, "gap_analysis": 0})
        if doc:
            scenarios.append(doc)

    return {"scenarios": scenarios}


@api_router.delete("/schedule/{schedule_id}")
async def delete_schedule(schedule_id: str):
    await db.schedules.delete_one({"id": schedule_id})
    return {"deleted": schedule_id}


@api_router.patch("/schedule/{schedule_id}")
async def rename_schedule(schedule_id: str, payload: dict = Body(...)):
    new_name = payload.get("name", "")
    if new_name:
        await db.schedules.update_one({"id": schedule_id}, {"$set": {"name": new_name}})
    return {"id": schedule_id, "name": new_name}


def parse_csv_to_dict(content: str) -> dict:
    import csv
    reader = csv.DictReader(io.StringIO(content))
    result = {d: {} for d in DAYS}
    for row in reader:
        interval = row.get("Interval", row.get("interval", "")).strip()
        if not interval:
            continue
        for day in DAYS:
            val = row.get(day, "0").strip()
            try:
                result[day][interval] = float(val) if val else 0
            except ValueError:
                result[day][interval] = 0
    return result


@api_router.get("/schedules")
async def list_schedules():
    schedules = await db.schedules.find(
        {}, {"_id": 0, "id": 1, "name": 1, "timestamp": 1, "summary": 1, "sla": 1, "off_day_profiles": 1}
    ).sort("timestamp", -1).to_list(50)
    return schedules


@api_router.get("/schedule/{schedule_id}")
async def get_schedule(schedule_id: str):
    doc = await db.schedules.find_one({"id": schedule_id}, {"_id": 0})
    if not doc:
        return {"error": "Schedule not found"}
    return doc


@api_router.get("/export/{schedule_id}/{export_type}")
async def export_xlsx(schedule_id: str, export_type: str):
    """Export schedule data as Excel (.xlsx)."""
    doc = await db.schedules.find_one({"id": schedule_id}, {"_id": 0})
    if not doc:
        return {"error": "Schedule not found"}

    output = io.BytesIO()
    wb = xlsxwriter.Workbook(output)
    header_fmt = wb.add_format({"bold": True, "bg_color": "#0033CC", "font_color": "#FFFFFF", "border": 1, "align": "center"})
    data_fmt = wb.add_format({"border": 1, "align": "center"})
    num_fmt = wb.add_format({"border": 1, "align": "center", "num_format": "0"})
    pct_fmt = wb.add_format({"border": 1, "align": "center", "num_format": "0.00\"%\""})

    if export_type == "shiftwise":
        ws = wb.add_worksheet("Shift Counts")
        headers = ["Shift ID", "Project", "Time Range"] + DAYS + ["Total"]
        for i, h in enumerate(headers):
            ws.write(0, i, h, header_fmt)
        for r_idx, row in enumerate(doc["shiftwise"]):
            ws.write(r_idx + 1, 0, row["shift_id"], data_fmt)
            ws.write(r_idx + 1, 1, row["project"], data_fmt)
            ws.write(r_idx + 1, 2, row["label"], data_fmt)
            for d_idx, d in enumerate(DAYS):
                ws.write(r_idx + 1, 3 + d_idx, row[d], num_fmt)
            ws.write(r_idx + 1, 10, row["total"], num_fmt)

    elif export_type == "roster":
        ws = wb.add_worksheet("Agent Roster")
        headers = ["Agent ID", "Off Days"] + DAYS
        for i, h in enumerate(headers):
            ws.write(0, i, h, header_fmt)
        for r_idx, row in enumerate(doc["roster"]):
            ws.write(r_idx + 1, 0, row["agent_id"], data_fmt)
            ws.write(r_idx + 1, 1, row["off_days"], data_fmt)
            for d_idx, d in enumerate(DAYS):
                ws.write(r_idx + 1, 2 + d_idx, row[d], data_fmt)

    elif export_type == "gap":
        ws = wb.add_worksheet("Gap Analysis")
        headers = ["Interval"]
        for d in DAYS:
            headers += [f"{d[:3]} Eng Req", f"{d[:3]} Eng Dep", f"{d[:3]} Eng Gap",
                        f"{d[:3]} Lang Req", f"{d[:3]} Lang Dep", f"{d[:3]} Lang Gap"]
        for i, h in enumerate(headers):
            ws.write(0, i, h, header_fmt)
        for r_idx, row in enumerate(doc["gap_analysis"]):
            ws.write(r_idx + 1, 0, row["interval"], data_fmt)
            col = 1
            for d in DAYS:
                ws.write(r_idx + 1, col, row[f"{d}_eng_req"], num_fmt)
                ws.write(r_idx + 1, col + 1, row[f"{d}_eng_deployed"], num_fmt)
                ws.write(r_idx + 1, col + 2, row[f"{d}_eng_gap"], num_fmt)
                ws.write(r_idx + 1, col + 3, row[f"{d}_lang_req"], num_fmt)
                ws.write(r_idx + 1, col + 4, row[f"{d}_lang_deployed"], num_fmt)
                ws.write(r_idx + 1, col + 5, row[f"{d}_lang_gap"], num_fmt)
                col += 6

    elif export_type == "sla":
        ws = wb.add_worksheet("SLA Analysis")
        headers = ["Day", "Eng Req", "Eng Met", "Eng SLA%", "Lang Req", "Lang Met", "Lang SLA%",
                    "Combined Req", "Combined Met", "Combined SLA%"]
        for i, h in enumerate(headers):
            ws.write(0, i, h, header_fmt)
        for r_idx, row in enumerate(doc["sla"]["daily"]):
            ws.write(r_idx + 1, 0, row["day"], data_fmt)
            ws.write(r_idx + 1, 1, row["english_required"], num_fmt)
            ws.write(r_idx + 1, 2, row["english_met"], num_fmt)
            ws.write(r_idx + 1, 3, row["english_sla"], pct_fmt)
            ws.write(r_idx + 1, 4, row["language_required"], num_fmt)
            ws.write(r_idx + 1, 5, row["language_met"], num_fmt)
            ws.write(r_idx + 1, 6, row["language_sla"], pct_fmt)
            ws.write(r_idx + 1, 7, row["combined_required"], num_fmt)
            ws.write(r_idx + 1, 8, row["combined_met"], num_fmt)
            ws.write(r_idx + 1, 9, row["combined_sla"], pct_fmt)

    wb.close()
    output.seek(0)
    return StreamingResponse(
        io.BytesIO(output.getvalue()),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={export_type}_{schedule_id[:8]}.xlsx"}
    )


app.include_router(api_router)

app.add_middleware(
    CORSMiddleware,
    allow_credentials=True,
    allow_origins=os.environ.get('CORS_ORIGINS', '*').split(','),
    allow_methods=["*"],
    allow_headers=["*"],
)

logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

@app.on_event("shutdown")
async def shutdown_db_client():
    client.close()
